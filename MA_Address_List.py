import os
import re
import io
import sys
import logging
from datetime import datetime, date



import requests
import pandas as pd
import pyodbc
from bs4 import BeautifulSoup
from dateutil.parser import parse as dateparse

# =========================
# Config (override via ENV)
# =========================
TARGET_PAGE = "https://www.mass.gov/lists/massachusetts-licensed-insurance-companies"
XLS_NAME_PATTERN = re.compile(r"Massachusetts\s+Licensed\s+Or\s+Approved\s+Companies\.xls", re.I)
today_str = datetime.today().strftime("%m%d%Y")


SQL_SERVER   = os.getenv("SQL_SERVER",   "AE1SQLWPV20")
SQL_DATABASE = os.getenv("SQL_DATABASE", "JiLi")
SQL_SCHEMA   = os.getenv("SQL_SCHEMA",   "dbo")
SQL_TABLE_BASE    = os.getenv("SQL_TABLE",    "address_list")
SQL_TABLE    = f"{SQL_TABLE_BASE}_{today_str}"

ODBC_DRIVER  = os.getenv("ODBC_DRIVER", "ODBC Driver 17 for SQL Server")  # or 18
TRUSTED_CONN = os.getenv("TRUSTED_CONN", "1") not in ("0", "false", "False")
SQL_USER     = os.getenv("SQL_USER")
SQL_PASSWORD = os.getenv("SQL_PASSWORD")

# =========
# Logging
# =========
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger("MA_Address_List")

# =========
# Helpers
# =========
def get_sql_connection():
    """Return a pyodbc connection with autocommit=True."""
    if TRUSTED_CONN:
        conn_str = (
            f"DRIVER={{{ODBC_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            "Trusted_Connection=yes"
        )
    else:
        if not SQL_USER or not SQL_PASSWORD:
            raise RuntimeError("SQL_USER/SQL_PASSWORD required for SQL Authentication")
        conn_str = (
            f"DRIVER={{{ODBC_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USER};PWD={SQL_PASSWORD}"
        )
    return pyodbc.connect(conn_str, autocommit=True)

def find_xls_url() -> str:
    """Find the 'Massachusetts Licensed Or Approved Companies.xls' link on the page."""
    log.info("Requesting Mass.gov listing page…")
    r = requests.get(TARGET_PAGE, timeout=60)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "lxml")

    for a in soup.find_all("a", href=True):
        text = (a.get_text() or "").strip()
        href = a["href"]
        if XLS_NAME_PATTERN.search(text) or XLS_NAME_PATTERN.search(href):
            if href.startswith("//"):
                href = "https:" + href
            elif href.startswith("/"):
                href = "https://www.mass.gov" + href
            log.info(f"Found XLS link: {text} -> {href}")
            return href

    raise RuntimeError("Could not find the 'Massachusetts Licensed Or Approved Companies.xls' link.")

def download_file(url: str) -> bytes:
    log.info("Downloading XLS/XLSX file…")
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content

def is_xlsx(file_bytes: bytes) -> bool:
    """XLSX (OOXML) are ZIP files and start with PK\x03\x04."""
    return file_bytes[:4] == b"PK\x03\x04"

def read_update_date_from_b4(xbytes: bytes):
    """
    Return a date (or None) from cell B4 of the first worksheet.
    Works for both .xlsx (openpyxl) and legacy .xls (xlrd==1.2.0).
    """
    if is_xlsx(xbytes):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(xbytes), data_only=True, read_only=True)
        ws = wb.worksheets[0]
        val = ws["B4"].value
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if val is not None:
            try: return dateparse(str(val)).date()
            except Exception: return None
        return None
    else:
        import xlrd  # must be 1.2.0 for .xls
        book = xlrd.open_workbook(file_contents=xbytes)
        sheet = book.sheet_by_index(0)
        cell = sheet.cell(3, 1)  # B4 (0-based)
        if cell.ctype == xlrd.XL_CELL_DATE:
            dt = datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
            return dt.date()
        try: return dateparse(str(cell.value)).date()
        except Exception: return None

def detect_header_row(df_raw: pd.DataFrame):
    """
    Heuristically find the header row index containing our expected column names.
    """
    expected = {"Company Type", "NAIC #", "Company", "Address", "City", "State", "Zip", "Phone"}
    scan_rows = min(40, len(df_raw))
    for idx in range(scan_rows):
        row_vals = set(str(x).strip() for x in df_raw.iloc[idx].tolist())
        if len(expected.intersection(row_vals)) >= 4:
            return idx
    return None

def load_table_dataframe(xbytes: bytes) -> pd.DataFrame:
    """Load the data table into a normalized DataFrame (handles .xls and .xlsx)."""
    engine = "openpyxl" if is_xlsx(xbytes) else "xlrd"

    # Pass 1: read without header to find header row
    df_raw = pd.read_excel(io.BytesIO(xbytes), header=None, engine=engine)
    hdr_idx = detect_header_row(df_raw)
    if hdr_idx is None:
        hdr_idx = 0

    # Pass 2: read with header row
    df = pd.read_excel(io.BytesIO(xbytes), header=hdr_idx, engine=engine)

    # Normalize headers
    rename_map = {
        "Company Type": "company_type",
        "NAIC #": "naic",
        "Company": "company",
        "Address": "address",
        "City": "city",
        "State": "state",
        "Zip": "zip",
        "Phone": "phone",
    }
    def norm_col(c):
        c0 = str(c).strip()
        for k, v in rename_map.items():
            if c0.lower() == k.lower():
                return v
        return c0

    df.columns = [norm_col(c) for c in df.columns]

    keep = ["company_type", "naic", "company", "address", "city", "state", "zip", "phone"]
    present = [c for c in keep if c in df.columns]
    df = df[present].copy()

    # Clean strings
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()

    # NAIC numeric-only
    if "naic" in df.columns:
        df["naic"] = df["naic"].str.extract(r"(\d+)", expand=False)

    # ZIP 5 or ZIP+4
    if "zip" in df.columns:
        df["zip"] = df["zip"].str.extract(r"(\d{5}(?:-\d{4})?)", expand=False)

    return df

def recreate_table(conn):
    ddl = f"""
    IF OBJECT_ID('{SQL_SCHEMA}.{SQL_TABLE}', 'U') IS NOT NULL
        DROP TABLE {SQL_SCHEMA}.{SQL_TABLE};

    CREATE TABLE {SQL_SCHEMA}.{SQL_TABLE}(
        company_type  VARCHAR(150)  NULL,
        naic          VARCHAR(20)   NULL,
        company       VARCHAR(255)  NULL,
        address       VARCHAR(255)  NULL,
        city          VARCHAR(120)  NULL,
        state         VARCHAR(10)   NULL,   -- was 2
        zip           VARCHAR(20)   NULL,
        phone         VARCHAR(40)   NULL,
        update_dt     DATE          NULL
    );
    """
    with conn.cursor() as cur:
        cur.execute(ddl)

def clean_and_trim(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Normalize state to 2-letter when possible
    if "state" in df.columns:
        st = df["state"].str.extract(r"([A-Za-z]{2})", expand=False).str.upper()
        df.loc[st.notna(), "state"] = st

    # Normalize ZIP
    if "zip" in df.columns:
        df["zip"] = df["zip"].str.extract(r"(\d{5}(?:-\d{4})?)", expand=False)

    # Keep digits in NAIC
    if "naic" in df.columns:
        df["naic"] = df["naic"].str.extract(r"(\d+)", expand=False)

    # Trim and clean
    trims = {
        "company_type": 150,
        "naic": 20,
        "company": 255,
        "address": 255,
        "city": 120,
        "state": 10,
        "zip": 20,
        "phone": 40,
    }
    for col, maxlen in trims.items():
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.slice(0, maxlen)
            # replace 'nan'/'NaN' strings with None
            df[col] = df[col].replace({"nan": None, "NaN": None, "None": None})

    return df



def insert_dataframe(conn, df: pd.DataFrame, update_dt_val):
    """Bulk insert rows; append update_dt if not present."""
    cols = ["company_type","naic","company","address","city","state","zip","phone","update_dt"]
    for c in cols:
        if c not in df.columns:
            df[c] = None

    if df.get("update_dt").isnull().all():
        df["update_dt"] = update_dt_val

    df = df[cols].copy()
    df = df.where(pd.notnull(df), None)
    df = df.replace({"nan": None, "NaN": None})

    placeholders = ", ".join(["?"] * len(cols))
    sql = f"INSERT INTO {SQL_SCHEMA}.{SQL_TABLE} ({', '.join(cols)}) VALUES ({placeholders})"

    with conn.cursor() as cur:
        cur.fast_executemany = True
        cur.executemany(sql, df.values.tolist())

    log.info(f"Inserted {len(df)} rows into {SQL_SCHEMA}.{SQL_TABLE}.")

# =====
# Main
# =====
def main():
    try:
        xls_url = find_xls_url()
        file_bytes = download_file(xls_url)

        update_dt = read_update_date_from_b4(file_bytes)
        if update_dt:
            log.info(f"Update date (B4): {update_dt.isoformat()}")
        else:
            log.warning("Could not read update date from B4; leaving update_dt as NULL.")

        df = load_table_dataframe(file_bytes)
        df = clean_and_trim(df) 
        # ensure column present even if None (insert_dataframe also protects)
        if "update_dt" not in df.columns:
            df["update_dt"] = update_dt

        conn = get_sql_connection()
        recreate_table(conn)
        insert_dataframe(conn, df, update_dt)

        log.info("All done ✅")
    except Exception as e:
        log.exception(f"Failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
