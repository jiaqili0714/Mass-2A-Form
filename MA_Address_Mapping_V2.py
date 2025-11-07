import os
import re
import io
import sys
import logging
from datetime import datetime, date

import requests
import pandas as pd
import pyodbc
from bs4 import BeautifulSoup
from dateutil.parser import parse as dateparse

# =========================
# Config (override via ENV)
# =========================
TARGET_PAGE = "https://www.mass.gov/lists/massachusetts-licensed-insurance-companies"
XLS_NAME_PATTERN = re.compile(r"Massachusetts\s+Licensed\s+Or\s+Approved\s+Companies\.xls", re.I)
today_str = datetime.today().strftime("%m%d%Y")

# --- Local File Archive (NEW) ---
# Folder to save the raw downloaded .xls/.xlsx file
ARCHIVE_FOLDER = os.getenv("ARCHIVE_FOLDER", r"\\njredbf2001\ProductManagement\Product\Auto\MASSACHUSETTS\Operational Processes\2A Form\MA Gov Company Address List")

# --- SQL Config ---
SQL_SERVER   = os.getenv("SQL_SERVER",   "ae1sqlwpv20")  # e.g. "AE1SQLWPV20"
SQL_DATABASE = os.getenv("SQL_DATABASE", "JiLi") # Target DB for all writes
SQL_SCHEMA   = os.getenv("SQL_SCHEMA",   "dbo")

# --- Part 2 (Mapping) ---
SQL_MAPPING_TABLE = "MA_2A_Form_Mapping"
RMV_SOURCE_DB = "CO1SQLWPV10_EnterpriseServices"   # Database where RMV_CARRIER_NAME table lives, if using NE server it's CO1SQLWPV10, if using AE1SQLWPV20 server it's CO1SQLWPV10_EnterpriseServices
RMV_SOURCE_TABLE = "EnterpriseServices.[dbo].[RMV_CARRIER_NAME]"

# --- Connection ---
ODBC_DRIVER  = os.getenv("ODBC_DRIVER", "ODBC Driver 17 for SQL Server")
# TRUSTED_CONN=1 uses Windows Auth (Trusted_Connection=yes)
TRUSTED_CONN = os.getenv("TRUSTED_CONN", "1") not in ("0", "false", "False")
SQL_USER     = os.getenv("SQL_USER")     # Ignored if TRUSTED_CONN=1
SQL_PASSWORD = os.getenv("SQL_PASSWORD") # Ignored if TRUSTED_CONN=1


# =========================
# Logging
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger("MA_RMV_Mapping")

# =========================
# Normalization Function
# =========================

# Define the stopwords from your SQL logic
STOPWORDS = {
    'INC', 'INCORPORATED', 'LLC', 'L.L.C', 'CO', 'COMPANY', 'CORP', 'CORPORATION',
    'GROUP', 'HOLDINGS', 'MUTUAL', 'ASSOCIATION', 'ASSN', 'ASSOCIATES',
    'INSURANCE', 'INS', 'CASUALTY', 'INDEMNITY', 'ASSURANCE',
    'FIRE', 'MARINE', 'PROPERTY', 'P&C', 'PC',
    'THE'
}

def normalize_name(s: str) -> str | None:
    """
    Translates the dbo.NormalizeInsName SQL function to Python.
    """
    if not s or pd.isna(s):
        return None

    x = str(s).upper()
    x = x.replace('&', ' AND ')
    x = re.sub(r'[.,\'"/\\()[\]{}:-]', ' ', x)
    x = re.sub(r'\s+', ' ', x).strip()
    
    if x.startswith('THE '):
        x = x[4:].lstrip()
        
    while len(x) > 0:
        last_space_idx = x.rfind(' ')
        if last_space_idx == -1:
            last_tok = x
        else:
            last_tok = x[last_space_idx + 1:]
            
        if last_tok in STOPWORDS:
            if last_space_idx == -1:
                x = ''
                break
            else:
                x = x[:last_space_idx].rstrip()
        else:
            break

    x = re.sub(r'\s+', ' ', x).strip()
    return x if x else None


# =========================
# Helpers
# =========================
def get_sql_connection():
    """Return a pyodbc connection with autocommit=True."""
    if TRUSTED_CONN:
        log.info("Connecting using Windows Authentication (Trusted_Connection=yes)")
        conn_str = (
            f"DRIVER={{{ODBC_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            "Trusted_Connection=yes"
        )
    else:
        if not SQL_USER or not SQL_PASSWORD:
            raise RuntimeError("SQL_USER/SQL_PASSWORD required for SQL Authentication")
        log.info(f"Connecting using SQL Authentication (User: {SQL_USER})")
        conn_str = (
            f"DRIVER={{{ODBC_DRIVER}}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USER};PWD={SQL_PASSWORD}"
        )
    return pyodbc.connect(conn_str, autocommit=True)

def find_xls_url() -> str:
    """Find the 'Massachusetts Licensed Or Approved Companies.xls' link on the page."""
    log.info(f"Requesting Mass.gov page: {TARGET_PAGE}")
    r = requests.get(TARGET_PAGE, timeout=60)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "lxml")

    for a in soup.find_all("a", href=True):
        text = (a.get_text() or "").strip()
        href = a["href"]
        if XLS_NAME_PATTERN.search(text) or XLS_NAME_PATTERN.search(href):
            if href.startswith("//"):
                href = "https:" + href
            elif href.startswith("/"):
                href = "https://www.mass.gov" + href
            log.info(f"Found XLS link: {text} -> {href}")
            return href

    raise RuntimeError("Could not find the 'Massachusetts Licensed Or Approved Companies.xls' link.")

def download_file(url: str) -> bytes:
    log.info(f"Downloading file from {url}...")
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content

def is_xlsx(file_bytes: bytes) -> bool:
    """XLSX (OOXML) are ZIP files and start with PK\x03\x04."""
    return file_bytes[:4] == b"PK\x03\x04"

def read_update_date_from_b4(xbytes: bytes):
    """Return a date (or None) from cell B4 of the first worksheet."""
    try:
        if is_xlsx(xbytes):
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(xbytes), data_only=True, read_only=True)
            ws = wb.worksheets[0]
            val = ws["B4"].value
        else:
            import xlrd  # must be 1.2.0 for .xls
            book = xlrd.open_workbook(file_contents=xbytes)
            sheet = book.sheet_by_index(0)
            cell = sheet.cell(3, 1)  # B4 (0-based)
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt = datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                return dt.date()
            val = cell.value

        if isinstance(val, datetime): return val.date()
        if isinstance(val, date): return val
        if val is not None:
            try: return dateparse(str(val)).date()
            except Exception: return None
            
    except Exception as e:
        log.warning(f"Could not read update date from B4: {e}")
        
    return None

def detect_header_row(df_raw: pd.DataFrame):
    """Heuristically find the header row index."""
    expected = {"Company Type", "NAIC #", "Company", "Address", "City", "State", "Zip", "Phone"}
    scan_rows = min(40, len(df_raw))
    for idx in range(scan_rows):
        row_vals = set(str(x).strip() for x in df_raw.iloc[idx].tolist())
        if len(expected.intersection(row_vals)) >= 4:
            return idx
    return None

def load_table_dataframe(xbytes: bytes) -> pd.DataFrame:
    """Load the data table into a normalized DataFrame."""
    engine = "openpyxl" if is_xlsx(xbytes) else "xlrd"
    
    df_raw = pd.read_excel(io.BytesIO(xbytes), header=None, engine=engine)
    hdr_idx = detect_header_row(df_raw)
    if hdr_idx is None:
        log.warning("Could not detect header row, defaulting to 0.")
        hdr_idx = 0
    log.info(f"Detected header row at index: {hdr_idx}")

    df = pd.read_excel(io.BytesIO(xbytes), header=hdr_idx, engine=engine)

    rename_map = {
        "Company Type": "company_type", "NAIC #": "naic", "Company": "company",
        "Address": "address", "City": "city", "State": "state", "Zip": "zip", "Phone": "phone",
    }
    def norm_col(c):
        c0 = str(c).strip()
        for k, v in rename_map.items():
            if c0.lower() == k.lower(): return v
        return c0.lower().replace(' ', '_')

    df.columns = [norm_col(c) for c in df.columns]

    keep = ["company_type", "naic", "company", "address", "city", "state", "zip", "phone"]
    present = [c for c in keep if c in df.columns]
    df = df[present].copy()
    
    return df

def clean_and_trim(df: pd.DataFrame) -> pd.DataFrame:
    """Cleans and trims the DataFrame."""
    df = df.copy()

    # Clean object/string columns
    for c in df.columns:
        if pd.api.types.is_object_dtype(df[c]):
            df[c] = df[c].astype(str).str.strip()
    
    # Normalize state (already string)
    if "state" in df.columns:
        st = df["state"].str.extract(r"([A-Za-z]{2})", expand=False).str.upper()
        df.loc[st.notna(), "state"] = st

    # Normalize ZIP (might be numeric)
    if "zip" in df.columns:
        # FIX: Convert to string first to handle numeric zips
        df["zip"] = df["zip"].astype(str).str.extract(r"(\d{5}(?:-\d{4})?)", expand=False)

    # Keep digits in NAIC (this was the error source)
    if "naic" in df.columns:
        # FIX: Convert to string *before* using .str accessor
        df["naic"] = df["naic"].astype(str).str.extract(r"(\d+)", expand=False)

    # Trim all columns to max lengths
    trims = {
        "company_type": 150, "naic": 20, "company": 255, "address": 255,
        "city": 120, "state": 10, "zip": 20, "phone": 40,
    }
    for col, maxlen in trims.items():
        if col in df.columns:
            # FIX: Ensure column is string before slicing
            df[col] = df[col].astype(str).str.slice(0, maxlen)
            
    # Replace all forms of null/nan with None
    df = df.replace({"nan": None, "NaN": None, "None": None, "NA": None, "<NA>": None})
    df = df.where(pd.notnull(df), None)

    return df

# --- Part 2 (Mapping) SQL Helpers ---

def get_rmv_data(conn) -> pd.DataFrame:
    """Pulls the RMV carrier list from the source DB."""
    query = f"""
    SELECT [CARRIER_NAME]
    FROM {RMV_SOURCE_DB}.{RMV_SOURCE_TABLE}
    WHERE [CARRIER_NAME] IS NOT NULL
    """
    log.info(f"Querying RMV carrier names from {RMV_SOURCE_DB}...")
    df_rmv = pd.read_sql_query(query, conn)
    log.info(f"Loaded {len(df_rmv)} rows from RMV table.")
    df_rmv = df_rmv.drop_duplicates(subset=['CARRIER_NAME']).reset_index(drop=True)
    log.info(f"Reduced to {len(df_rmv)} unique RMV names.")
    return df_rmv
    
def apply_hardcoded_matches(df_rmv: pd.DataFrame) -> pd.DataFrame:
    """
    Applies the custom override logic to map specific RMV names
    to their known Mass Gov equivalents *before* normalization.
    """
    log.info("Applying hardcoded match logic...")
    # Default: the match target is the original name
    df_rmv['rmv_match_target'] = df_rmv['CARRIER_NAME']
    
    # 1. Pattern Match: XXXX(Pilgrim) -> Pilgrim Insurance Company
    pilgrim_mask = df_rmv['CARRIER_NAME'].str.contains(r'\(Pilgrim\)', na=False, case=False)
    df_rmv.loc[pilgrim_mask, 'rmv_match_target'] = 'Pilgrim Insurance Company'
    log.info(f"Mapped {pilgrim_mask.sum()} RMV names to 'Pilgrim Insurance Company'")

    # 2. Exact Name Overrides
    overrides = {
        # Original overrides
        'Privilege Underwriters Reciprocal Exchange (PURE)': 'Privilege Underwriters Reciprocal Exchange',
        'Metropolitan Property and Casualty Insurance Company': 'Farmers Casualty Insurance Company',
        'Electric Insurance Company': 'Plymouth Rock Assurance Corporation',
    
        'Foremost Insurance Company': 'Foremost Property and Casualty Insurance Company',
        'Citation Insurance Company, MA': 'Citation Insurance Company',
        'IDS Property Casualty Insurance Company': 'American Family Connect Insurance Company',
        'Seaworthy Insurance Company': 'GEICO Marine Insurance Company'
    }
    
    for rmv_name, target_name in overrides.items():
        override_mask = (df_rmv['CARRIER_NAME'] == rmv_name)
        df_rmv.loc[override_mask, 'rmv_match_target'] = target_name
        # Log count for each override
        count = override_mask.sum()
        if count > 0:
            log.info(f"Mapped {count} RMV names from '{rmv_name}' to '{target_name}'")
        
    return df_rmv

def recreate_mapping_table(conn):
    """Drops and recreates the final mapping table."""
    ddl = f"""
    IF OBJECT_ID('{SQL_SCHEMA}.{SQL_MAPPING_TABLE}', 'U') IS NOT NULL
        DROP TABLE {SQL_SCHEMA}.{SQL_MAPPING_TABLE};

    CREATE TABLE {SQL_SCHEMA}.{SQL_MAPPING_TABLE}(
        rmv_name        VARCHAR(255) NULL,
        mass_gov_name   VARCHAR(255) NULL,
        address         VARCHAR(255) NULL,
        city            VARCHAR(120) NULL,
        state           VARCHAR(10)  NULL,
        zip             VARCHAR(20)  NULL,
        phone           VARCHAR(40)  NULL,
        update_dt       DATE         NULL
    );
    """
    with conn.cursor() as cur:
        log.info(f"Recreating mapping table: {SQL_SCHEMA}.{SQL_MAPPING_TABLE}")
        cur.execute(ddl)
        
def insert_mapping_dataframe(conn, df: pd.DataFrame):
    """Bulk insert rows into the final mapping table."""
    cols = ["rmv_name", "mass_gov_name", "address", "city", "state", "zip", "phone", "update_dt"]
    df_insert = df[cols].copy()
    
    df_insert = df_insert.where(pd.notnull(df_insert), None)

    placeholders = ", ".join(["?"] * len(cols))
    sql = f"INSERT INTO {SQL_SCHEMA}.{SQL_MAPPING_TABLE} ({', '.join(cols)}) VALUES ({placeholders})"

    with conn.cursor() as cur:
        cur.fast_executemany = True
        cur.executemany(sql, df_insert.values.tolist())

    log.info(f"Inserted {len(df_insert)} rows into {SQL_SCHEMA}.{SQL_MAPPING_TABLE}.")


# =========================
# Main Execution
# =========================
def main():
    try:
        conn = get_sql_connection()
        log.info(f"Connected to SQL Server: {SQL_SERVER}, DB: {SQL_DATABASE}")
        
        # --- PART 1: Download, Clean, and Archive Mass Gov List ---
        log.info("--- Starting Part 1: Mass Gov Download & Archive ---")
        xls_url = find_xls_url()
        file_bytes = download_file(xls_url)
        
        # --- 1.1 Save raw file to archive folder ---
        os.makedirs(ARCHIVE_FOLDER, exist_ok=True)
        file_ext = ".xlsx" if is_xlsx(file_bytes) else ".xls"
        archive_filename = f"MA_Licensed_Companies_{date.today().strftime('%Y%m%d')}{file_ext}"
        archive_path = os.path.join(ARCHIVE_FOLDER, archive_filename)
        
        with open(archive_path, 'wb') as f:
            f.write(file_bytes)
        log.info(f"Raw file saved for record at {archive_path}")

        # --- 1.2 Load data into DataFrame for Part 2 ---
        update_dt = read_update_date_from_b4(file_bytes)
        if update_dt:
            log.info(f"Update date (from B4): {update_dt.isoformat()}")
        else:
            log.warning("Could not read update date from B4.")

        df_mass_gov_raw = load_table_dataframe(file_bytes)
        df_mass_gov_cleaned = clean_and_trim(df_mass_gov_raw)
        
        # --- 1.3 Filter for 'Property & Casualty' ---
        log.info(f"Loaded {len(df_mass_gov_cleaned)} total rows from Mass Gov list.")
        filter_mask = df_mass_gov_cleaned['company_type'].str.contains(
            'Property & Casualty', 
            case=False, 
            na=False
        )
        df_mass_gov = df_mass_gov_cleaned[filter_mask].copy()
        
        if len(df_mass_gov) == 0:
            log.warning("Filter 'Property & Casualty' resulted in 0 companies. Check the string.")
        
        log.info(f"--- Part 1: Download & Archive Complete. {len(df_mass_gov)} 'P&C' rows loaded for processing. ---")

        
        # --- PART 2: Load RMV, Match (Multi-Pass), and Save Mapping Table ---
        log.info("--- Starting Part 2: RMV Mapping ---")
        
        # 2.1 Load RMV Data
        df_rmv_raw = get_rmv_data(conn)
        
        # --- 2.2 Pass 1: Exact Raw Match ---
        log.info("--- Starting Pass 1: Exact Raw Match ---")
        df_exact_matches = pd.merge(
            df_rmv_raw,
            df_mass_gov,
            left_on='CARRIER_NAME',
            right_on='company',
            how='inner',
            suffixes=('_rmv', '_pass1')
        )
        log.info(f"Found {len(df_exact_matches)} exact raw matches (Pass 1).")
        
        # Get list of RMV names that are now matched
        matched_rmv_names = df_exact_matches['CARRIER_NAME'].unique()

        # --- 2.3 Pass 2: Normalized Match (for unmatched) ---
        log.info("--- Starting Pass 2: Normalized Match ---")
        
        # Get RMV names that *did not* find an exact match
        df_rmv_unmatched = df_rmv_raw[~df_rmv_raw['CARRIER_NAME'].isin(matched_rmv_names)].copy()
        log.info(f"{len(df_rmv_unmatched)} RMV names remaining for normalization.")

        if not df_rmv_unmatched.empty:
            # 2.3a Apply Overrides (hardcodes)
            df_rmv_unmatched = apply_hardcoded_matches(df_rmv_unmatched)

            # 2.3b Normalize both sides
            log.info("Normalizing remaining names...")
            df_rmv_unmatched['normalized_name'] = df_rmv_unmatched['rmv_match_target'].apply(normalize_name)
            df_mass_gov['normalized_name'] = df_mass_gov['company'].apply(normalize_name)
            
            # Prep for merge (drop nulls)
            df_rmv_norm = df_rmv_unmatched.dropna(subset=['normalized_name', 'CARRIER_NAME'])
            df_mass_norm = df_mass_gov.dropna(subset=['normalized_name', 'company'])
            
            # 2.3c Perform normalized merge
            log.info("Performing exact match on normalized names...")
            df_normalized_matches = pd.merge(
                df_rmv_norm,
                df_mass_norm,
                on='normalized_name',
                how='inner',
                suffixes=('_rmv', '_pass2')
            )
            log.info(f"Found {len(df_normalized_matches)} normalized matches (Pass 2).")
        else:
            log.info("No RMV names left for normalized matching.")
            # Create empty DataFrame to avoid errors later
            df_normalized_matches = pd.DataFrame() 

        # --- 2.4 Combine and Construct Final Table ---
        log.info("Constructing final mapping table...")
        
        # Define the columns we want in the final table
        final_cols = ['CARRIER_NAME', 'company', 'address', 'phone', 'state', 'city', 'zip']
        
        # Format Pass 1 results
        df_final_pass1 = df_exact_matches[final_cols].copy()
        
        # Format Pass 2 results (if any)
        if not df_normalized_matches.empty:
            df_final_pass2 = df_normalized_matches[final_cols].copy()
        else:
            df_final_pass2 = pd.DataFrame(columns=final_cols) # Empty df with correct columns
            
        # Combine the results (Pass 1 first)
        df_mapping_combined = pd.concat([df_final_pass1, df_final_pass2], ignore_index=True)
        log.info(f"Total matches (Pass 1 + Pass 2): {len(df_mapping_combined)}")

        # Rename columns
        df_mapping_combined.rename(columns={
            'CARRIER_NAME': 'rmv_name',
            'company': 'mass_gov_name'
        }, inplace=True)
        
        # Add update_dt
        df_mapping_combined['update_dt'] = date.today()
        
        # Critical: Drop duplicates *after* combining.
        # By using keep='first', we prioritize the Pass 1 (exact) matches.
        df_mapping_final = df_mapping_combined.drop_duplicates(subset=['rmv_name'], keep='first')
        log.info(f"Final mapping table has {len(df_mapping_final)} unique RMV mappings.")

        # --- Override naming rule for (Pilgrim) rows ---
        # Keep the original RMV name as mass_gov_name, but retain Pilgrim's address info.
        pilgrim_mask_final = df_mapping_final['rmv_name'].str.contains(r'\(Pilgrim\)', case=False, na=False)
        df_mapping_final.loc[pilgrim_mask_final, 'mass_gov_name'] = df_mapping_final.loc[pilgrim_mask_final, 'rmv_name']
        log.info(f"Adjusted {pilgrim_mask_final.sum()} '(Pilgrim)' rows to keep RMV name as Mass Gov name while retaining Pilgrim address.")

        

        # --- 2.5 Save to SQL ---
        recreate_mapping_table(conn)
        insert_mapping_dataframe(conn, df_mapping_final)
        log.info("--- Part 2: RMV Mapping Complete ---")

        log.info("All done ✅")

    except Exception as e:
        log.exception(f"Process Failed: {e}")
        sys.exit(1)
    finally:
        if 'conn' in locals() and conn:
            conn.close()
            log.info("SQL Connection closed.")

if __name__ == "__main__":
    # --- Dependencies needed ---
    # pip install pandas requests pyodbc beautifulsoup4 lxml openpyxl xlrd==1.2.0 python-dateutil
    main()
